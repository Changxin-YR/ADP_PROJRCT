from __future__ import annotations

import argparse
import gc
import json
import threading
import time
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, TypeVar
from uuid import uuid4

import psutil
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from backend.config.settings import Settings
from backend.layers.common.db.connection import get_connection
from backend.layers.features.data_exchange.data_exchange_service import DataExchangeService
from backend.layers.features.data_exchange.data_exchange_store import MySqlDataExchangeStore
from backend.scripts.readiness.common import evidence, require_disposable_database_name
from backend.scripts.readiness.database_drill import SEED_FILES
from backend.scripts.readiness.mysql_tools import (
    apply_migrations,
    apply_sql,
    execute_sql,
    require_database_tools,
)


T = TypeVar("T")
ROW_TARGET = 100_000
MAX_SECONDS_PER_FORMAT = 60
MAX_RSS_DELTA_BYTES = 512 * 1024 * 1024

def _settings(tools: Any, database: str) -> Settings:
    return Settings.from_env(
        {
            "APP_ENV": "test",
            "FLASK_SECRET_KEY": "readiness-export-flask-secret",
            "CSRF_SECRET_KEY": "readiness-export-csrf-secret",
            "MYSQL_HOST": str(tools.connection["host"]),
            "MYSQL_PORT": str(tools.connection["port"]),
            "MYSQL_DATABASE": database,
            "MYSQL_USER": str(tools.connection["user"]),
            "MYSQL_PASSWORD": str(tools.connection["password"]),
            "SESSION_COOKIE_SECURE": "false",
        }
    )


def _measure_peak_rss(operation: Callable[[], T]) -> tuple[T, float, int, int]:
    gc.collect()
    process = psutil.Process()
    baseline = process.memory_info().rss
    peak = baseline
    stop = threading.Event()

    def sample() -> None:
        nonlocal peak
        while not stop.wait(0.02):
            peak = max(peak, process.memory_info().rss)

    sampler = threading.Thread(target=sample, name="readiness-rss-sampler", daemon=True)
    sampler.start()
    started = time.perf_counter()
    try:
        result = operation()
    finally:
        elapsed = time.perf_counter() - started
        peak = max(peak, process.memory_info().rss)
        stop.set()
        sampler.join(timeout=2)
    return result, elapsed, baseline, peak


def _seed_rows(settings: Settings) -> tuple[int, int]:
    with get_connection(settings) as connection, connection.cursor() as cursor:
        cursor.execute(
            "INSERT INTO users(phone,login_name,name,password_hash,status) "
            "VALUES ('13999999991','TEST_EXPORT_USER','TEST_EXPORT_导出员','not-used','active')"
        )
        user_id = int(cursor.lastrowid)
        cursor.execute("SELECT id FROM organizations WHERE code='default'")
        organization_id = int(cursor.fetchone()["id"])
        cursor.execute(
            "INSERT INTO organizations(code,name,status) VALUES "
            "('TEST_EXPORT_SENTINEL_ORG','TEST_EXPORT_越权哨兵企业','active')"
        )
        sentinel_organization_id = int(cursor.lastrowid)

        insert_sql = (
            "INSERT INTO business_partners"
            "(organization_id,partner_type,code,name,contact_name,phone,status,created_by) "
            "VALUES (%s,'customer',%s,%s,%s,%s,'verified',%s)"
        )
        for start in range(0, ROW_TARGET, 2000):
            rows = [
                (
                    organization_id,
                    f"TEST_EXPORT_{index:06d}",
                    f"TEST_EXPORT_客户{index:06d}",
                    f"联系人{index:06d}",
                    f"13{index:09d}"[-11:],
                    user_id,
                )
                for index in range(start + 1, min(start + 2000, ROW_TARGET) + 1)
            ]
            cursor.executemany(insert_sql, rows)
        cursor.execute(
            insert_sql,
            (
                organization_id,
                "TEST_EXPORT_DRAFT_EXCLUDED",
                "TEST_EXPORT_草稿排除",
                "排除",
                "13000000001",
                user_id,
            ),
        )
        cursor.execute(
            "UPDATE business_partners SET status='draft' WHERE organization_id=%s AND code='TEST_EXPORT_DRAFT_EXCLUDED'",
            (organization_id,),
        )
        cursor.execute(
            insert_sql,
            (
                organization_id,
                "OTHER_PREFIX_EXCLUDED",
                "非搜索结果",
                "排除",
                "13000000002",
                user_id,
            ),
        )
        cursor.execute(
            insert_sql,
            (
                sentinel_organization_id,
                "TEST_EXPORT_SENTINEL_OUT_OF_SCOPE",
                "TEST_EXPORT_越权哨兵",
                "越权",
                "13000000003",
                user_id,
            ),
        )
    return organization_id, user_id


def _validate_xlsx(content: bytes) -> dict[str, Any]:
    started = time.perf_counter()
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["导出数据"]
    iterator = sheet.iter_rows(values_only=True)
    header = tuple(next(iterator))
    code_index = header.index("code")
    row_count = 0
    sentinel_found = False
    first_code = ""
    last_code = ""
    for values in iterator:
        code = str(values[code_index])
        row_count += 1
        first_code = first_code or code
        last_code = code
        sentinel_found = sentinel_found or "SENTINEL" in code
    guide = {str(key): str(value) for key, value in workbook["导出说明"].iter_rows(values_only=True)}
    workbook.close()
    return {
        "row_count": row_count,
        "metadata_row_count": int(guide["row_count"]),
        "request_id": guide["request_id"],
        "first_code": first_code,
        "last_code": last_code,
        "sentinel_found": sentinel_found,
        "validation_seconds": round(time.perf_counter() - started, 3),
    }


def _validate_pdf(content: bytes) -> dict[str, Any]:
    started = time.perf_counter()
    reader = PdfReader(BytesIO(content))
    metadata = reader.metadata
    title = str(metadata.title or "") if metadata else ""
    subject = str(metadata.subject or "") if metadata else ""
    return {
        "pages": len(reader.pages),
        "header_present": "customers" in title,
        "metadata_row_count_present": f"row_count={ROW_TARGET}" in subject,
        "sentinel_bytes_present": b"TEST_EXPORT_SENTINEL_OUT_OF_SCOPE" in content,
        "validation_seconds": round(time.perf_counter() - started, 3),
    }


def run_probe(output: Path) -> dict[str, Any]:
    tools = require_database_tools()
    database = require_disposable_database_name(f"adp_readiness_export_{uuid4().hex[:12]}")
    metrics: dict[str, Any] = {"database": database, "target_rows": ROW_TARGET}
    created = False
    cleanup_errors: list[str] = []
    passed = False
    try:
        execute_sql(tools, f"CREATE DATABASE `{database}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        created = True
        apply_migrations(tools, database)
        for seed in SEED_FILES:
            apply_sql(tools, database, seed)
        settings = _settings(tools, database)
        seed_started = time.perf_counter()
        organization_id, user_id = _seed_rows(settings)
        metrics["seed_seconds"] = round(time.perf_counter() - seed_started, 3)

        user = {
            "id": user_id,
            "name": "TEST_EXPORT_导出员",
            "permissions": ["data_exchange.export"],
            "roles": [{"code": "super_admin"}],
            "data_scopes": [{"scope_type": "farm", "organization_id": None}],
        }
        service = DataExchangeService(MySqlDataExchangeStore(settings), output.parent / "unused-attachments")
        format_metrics: dict[str, Any] = {}
        contents: dict[str, bytes] = {}
        for file_format in ("xlsx", "pdf"):
            request_id = f"readiness-{file_format}-{uuid4().hex[:12]}"
            (content, export_id), elapsed, baseline, peak = _measure_peak_rss(
                lambda current_format=file_format, current_request=request_id: service.export(
                    user,
                    organization_id=organization_id,
                    resource="customers",
                    file_format=current_format,
                    filters={"status": "verified", "search": "TEST_EXPORT_"},
                    request_id=current_request,
                )
            )
            contents[file_format] = content
            format_metrics[file_format] = {
                "seconds": round(elapsed, 3),
                "rss_baseline_bytes": baseline,
                "rss_peak_bytes": peak,
                "rss_delta_bytes": max(0, peak - baseline),
                "content_bytes": len(content),
                "export_id": export_id,
                "request_id": request_id,
            }

        format_metrics["xlsx"]["validation"] = _validate_xlsx(contents["xlsx"])
        format_metrics["pdf"]["validation"] = _validate_pdf(contents["pdf"])
        with get_connection(settings) as connection, connection.cursor() as cursor:
            cursor.execute(
                "SELECT format,row_count,request_id FROM data_export_audits WHERE exported_by=%s ORDER BY id",
                (user_id,),
            )
            audits = [dict(row) for row in cursor.fetchall()]
        metrics["formats"] = format_metrics
        metrics["audits"] = audits
        passed = (
            all(item["seconds"] <= MAX_SECONDS_PER_FORMAT for item in format_metrics.values())
            and all(item["rss_delta_bytes"] <= MAX_RSS_DELTA_BYTES for item in format_metrics.values())
            and format_metrics["xlsx"]["validation"]["row_count"] == ROW_TARGET
            and format_metrics["xlsx"]["validation"]["metadata_row_count"] == ROW_TARGET
            and not format_metrics["xlsx"]["validation"]["sentinel_found"]
            and format_metrics["pdf"]["validation"]["header_present"]
            and format_metrics["pdf"]["validation"]["metadata_row_count_present"]
            and not format_metrics["pdf"]["validation"]["sentinel_bytes_present"]
            and len(audits) == 2
            and all(int(item["row_count"]) == ROW_TARGET for item in audits)
        )
    except Exception as exc:
        metrics["error_type"] = type(exc).__name__
        metrics["error"] = str(exc)
    finally:
        if created:
            try:
                require_disposable_database_name(database)
                execute_sql(tools, f"DROP DATABASE IF EXISTS `{database}`")
            except Exception as exc:
                cleanup_errors.append(f"{database}: {exc}")
        metrics["cleanup_errors"] = cleanup_errors
        metrics["cleanup"] = not cleanup_errors
        passed = passed and not cleanup_errors

    result = evidence("large-export", metrics, passed=passed)
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(output.suffix + ".tmp")
    temporary.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    temporary.replace(output)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="ADP 100,000-row export readiness probe")
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    result = run_probe(args.output.resolve())
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0 if result["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())

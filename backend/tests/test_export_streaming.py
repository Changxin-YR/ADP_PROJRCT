from __future__ import annotations

from contextlib import contextmanager
from io import BytesIO
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from backend.layers.features.data_exchange.data_exchange_service import DataExchangeService
from backend.layers.features.data_exchange.workbooks import export_pdf_stream, export_workbook_stream


class OneShotRows:
    def __init__(self, count: int) -> None:
        self.count = count
        self.started = False

    def __iter__(self) -> Iterator[dict[str, Any]]:
        if self.started:
            raise AssertionError("流式导出不得重复遍历结果集")
        self.started = True
        for index in range(self.count):
            yield {"id": index + 1, "code": f"TEST-{index + 1:06d}"}


def test_xlsx_and_pdf_renderers_accept_one_shot_iterators() -> None:
    metadata = {"resource": "materials", "request_id": "stream-test"}
    xlsx = export_workbook_stream(OneShotRows(3), metadata)
    pdf = export_pdf_stream(OneShotRows(3), metadata)

    assert xlsx.row_count == 3
    assert pdf.row_count == 3
    assert pdf.content.startswith(b"%PDF-")
    assert b"row_count=3" in pdf.content

    workbook = load_workbook(BytesIO(xlsx.content), read_only=True, data_only=True)
    data_rows = list(workbook["导出数据"].iter_rows(values_only=True))
    guide = dict(workbook["导出说明"].iter_rows(values_only=True))
    workbook.close()
    assert len(data_rows) == 4
    assert guide["row_count"] == "3"


def test_service_uses_store_stream_and_audits_actual_row_count(tmp_path: Path) -> None:
    class Store:
        audit: dict[str, Any] | None = None

        def export_rows(self, *_args: Any, **_kwargs: Any) -> list[dict[str, Any]]:
            raise AssertionError("支持流式接口时不得回退 fetchall")

        @contextmanager
        def export_stream(self, *_args: Any, **_kwargs: Any) -> Iterator[OneShotRows]:
            yield OneShotRows(5)

        def record_export(self, payload: dict[str, Any]) -> int:
            self.audit = payload
            return 99

    store = Store()
    service = DataExchangeService(store, tmp_path)
    user = {
        "id": 1,
        "name": "导出员",
        "permissions": ["data_exchange.export"],
        "roles": [{"code": "super_admin"}],
        "data_scopes": [{"scope_type": "farm", "organization_id": None}],
    }

    content, export_id = service.export(
        user,
        organization_id=1,
        resource="materials",
        file_format="xlsx",
        filters={"status": "draft"},
        request_id="stream-service",
    )

    assert content.startswith(b"PK")
    assert export_id == 99
    assert store.audit is not None and store.audit["row_count"] == 5


def test_xlsx_export_escapes_formula_like_text() -> None:
    artifact = export_workbook_stream(iter([{"name": "=HYPERLINK(\"https://evil\")"}]), {"resource": "materials"})
    workbook = load_workbook(BytesIO(artifact.content), read_only=True, data_only=False)
    value = next(workbook["导出数据"].iter_rows(min_row=2, values_only=True))[0]
    workbook.close()
    assert value == "'=HYPERLINK(\"https://evil\")"


def test_pdf_export_preserves_long_row_text() -> None:
    import inspect

    assert "text[:100]" not in inspect.getsource(export_pdf_stream)

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from backend.app import create_app
from backend.tests.fake_auth_store import FakeAuthStore
from backend.tests.test_data_exchange import FakeExchangeStore, client, settings, token, upload, workbook


ROOT = Path(__file__).parents[2]

def test_all_templates_are_importable_except_inventory_ledger(tmp_path: Path) -> None:
    browser = client(tmp_path)
    templates = browser.get("/api/v1/data-exchange/templates").get_json()["data"]["items"]
    by_code = {item["code"]: item for item in templates}
    assert by_code["inventory-ledger"]["importable"] is False
    assert len(templates) >= 30
    for code, item in by_code.items():
        if code != "inventory-ledger":
            assert item["importable"] is True, f"{code} 应开放导入"


def test_inventory_ledger_import_is_rejected_with_explanation(tmp_path: Path) -> None:
    browser = client(tmp_path)
    response = browser.post(
        "/api/v1/data-exchange/imports/preview",
        data={"organization_id": "1", "template_code": "inventory-ledger", "file": (BytesIO(workbook([["LED-001", 1, "receipt", 10, "2026-08-17"]])), "ledger.xlsx")},
        headers=token(browser), content_type="multipart/form-data",
    )
    assert response.status_code == 409
    body = response.get_json()
    assert body["code"] == "IMPORT_TEMPLATE_NOT_IMPLEMENTED"
    assert "只追加账本" in body["message"]


def test_revoke_deletes_imported_drafts_and_rejects_repeat(tmp_path: Path) -> None:
    store = FakeExchangeStore()
    browser = client(tmp_path, store)
    preview = upload(browser, workbook([["MAT-REV-1", "撤销饲料", "饲料", "kg"]])).get_json()["data"]["batch"]
    confirmed = browser.post(f"/api/v1/data-exchange/imports/{preview['id']}/confirm", json={}, headers=token(browser))
    assert confirmed.status_code == 200
    assert len(store.materials) == 1

    revoked = browser.post(f"/api/v1/data-exchange/imports/{preview['id']}/revoke", json={}, headers=token(browser))
    assert revoked.status_code == 200
    assert revoked.get_json()["data"]["batch"]["status"] == "undone"
    assert store.materials == []

    again = browser.post(f"/api/v1/data-exchange/imports/{preview['id']}/revoke", json={}, headers=token(browser))
    assert again.status_code == 409
    assert again.get_json()["code"] == "IMPORT_NOT_IMPORTED"


def test_revoke_requires_import_permission_and_rejects_non_imported(tmp_path: Path) -> None:
    store = FakeExchangeStore()
    browser = client(tmp_path, store)
    preview = upload(browser, workbook([["MAT-REV-2", "待撤销饲料", "饲料", "kg"]])).get_json()["data"]["batch"]

    denied = browser.post(f"/api/v1/data-exchange/imports/{preview['id']}/revoke", json={}, headers=token(browser))
    assert denied.status_code == 409
    assert denied.get_json()["code"] == "IMPORT_NOT_IMPORTED"

    auth = FakeAuthStore()
    viewer = auth.add_user(phone="13800000809", login_name="view-only", password="Correct9!", status="active")
    viewer.update(permissions=["data_exchange.view"])
    app = create_app(settings(tmp_path), store=auth, data_exchange_store=store)
    restricted = app.test_client()
    csrf = restricted.get("/api/v1/auth/csrf").get_json()["data"]["csrf_token"]
    assert restricted.post("/api/v1/auth/login", json={"identifier": "view-only", "password": "Correct9!"}, headers={"X-CSRF-Token": csrf}).status_code == 200
    csrf = restricted.get("/api/v1/auth/csrf").get_json()["data"]["csrf_token"]
    forbidden = restricted.post(f"/api/v1/data-exchange/imports/{preview['id']}/revoke", json={}, headers={"X-CSRF-Token": csrf})
    assert forbidden.status_code == 403
    assert forbidden.get_json()["code"] == "FORBIDDEN"


def test_export_rejects_unknown_resource_and_keeps_metadata(tmp_path: Path) -> None:
    store = FakeExchangeStore()
    browser = client(tmp_path, store)
    unknown = browser.post("/api/v1/data-exchange/exports", json={"organization_id": 1, "resource": "not-a-resource", "format": "xlsx", "filters": {}}, headers=token(browser))
    assert unknown.status_code == 400
    assert unknown.get_json()["code"] == "EXPORT_RESOURCE_INVALID"
    response = browser.post(
        "/api/v1/data-exchange/exports", json={"organization_id": 1, "resource": "imports", "format": "xlsx", "filters": {"status": "ready"}},
        headers={**token(browser), "X-Request-ID": "meta-check"},
    )
    assert response.status_code == 200 and response.data.startswith(b"PK")
    assert store.exports[-1]["row_count"] == 0
    assert store.exports[-1]["request_id"] == "meta-check"
    # BUG-011：xlsx 导出文件内写入生成时间/导出人/筛选/口径等元数据。
    from openpyxl import load_workbook
    book = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
    meta = {str(row[0]): str(row[1]) for row in book["导出说明"].iter_rows(values_only=True) if row[0] is not None}
    assert meta["resource"] == "imports"
    assert meta["actor"] == "测试用户"
    assert meta["row_count"] == "0"
    assert meta["request_id"] == "meta-check"
    assert meta["generated_at"]
    assert "status" in meta["filters"]


def test_data_exchange_migration_protects_audit_and_import_facts() -> None:
    migration = ROOT / "database/migrations/017_data_exchange.sql"
    assert migration.exists(), "data exchange schema is not implemented"
    sql = migration.read_text(encoding="utf-8")
    for marker in ("CREATE TABLE data_import_batches", "UNIQUE KEY uq_data_import_file", "CREATE TABLE data_export_audits", "data_export_audits_no_update", "data_export_audits_no_delete"):
        assert marker in sql


def test_follow_up_migration_revokes_breed_worker_data_exchange_overgrant() -> None:
    migration = ROOT / "database/migrations/027_revoke_breed_worker_data_exchange.sql"
    assert migration.exists()
    sql = migration.read_text(encoding="utf-8").lower()
    assert "delete rp" in sql
    assert "r.code='breed_worker'" in sql
    assert "p.code in ('data_exchange.view','data_exchange.export')" in sql

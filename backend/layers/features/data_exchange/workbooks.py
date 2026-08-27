from __future__ import annotations

from datetime import date, datetime
from dataclasses import dataclass
from io import BytesIO
from itertools import chain
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from textwrap import wrap

from backend.layers.common.governance.lifecycle import DomainError
from backend.layers.features.data_exchange.template_catalog import Template


MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000


@dataclass(frozen=True)
class ExportArtifact:
    content: bytes
    row_count: int


def template_workbook(template: Template) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "数据"
    sheet.append([field.key for field in template.fields])
    sheet.append([field.example for field in template.fields])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="315C52")
    guide = book.create_sheet("填写说明")
    guide.append(["字段", "含义", "必填", "格式", "允许值", "示例"])
    for field in template.fields:
        guide.append([field.key, field.label, "是" if field.required else "否", field.kind, "/".join(field.allowed), field.example])
    guide.append(["模板版本", template.version, "", "", "", ""])
    return _save(book)


def preview_workbook(template: Template, content: bytes) -> dict[str, Any]:
    if not content or len(content) > MAX_IMPORT_BYTES:
        raise DomainError("IMPORT_FILE_SIZE_INVALID", "文件必须在 1 字节到 10 MB 之间", 400)
    try:
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
    except Exception as exc:
        raise DomainError("IMPORT_FILE_INVALID", "文件不是有效的 Excel 工作簿", 400) from exc
    values = sheet.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(values, ())]
    missing = [field.key for field in template.fields if field.required and field.key not in header]
    if missing:
        raise DomainError("IMPORT_COLUMNS_MISSING", f"缺少模板列：{', '.join(missing)}", 422)
    rows, errors, row_numbers = [], [], []
    for row_number, values_row in enumerate(values, start=2):
        if row_number > MAX_IMPORT_ROWS + 1:
            raise DomainError("IMPORT_ROWS_EXCEEDED", "单次导入不能超过 5000 行", 422)
        raw = dict(zip(header, values_row))
        if not any(value not in (None, "") for value in raw.values()):
            continue
        normalized: dict[str, Any] = {}
        for field in template.fields:
            value = raw.get(field.key)
            normalized[field.key] = _json_value(value)
            message = _validate(field.required, field.kind, field.allowed, value)
            if message:
                errors.append({"row": row_number, "column": field.key, "message": message, "value": _json_value(value)})
        rows.append(normalized)
        row_numbers.append(row_number)
    if not rows:
        raise DomainError("IMPORT_ROWS_EMPTY", "Excel 中没有可导入的数据行", 422)
    bad_rows = {item["row"] for item in errors}
    return {"preview_rows": rows, "row_numbers": row_numbers, "errors": errors, "total_rows": len(rows), "passed_rows": len(rows) - len(bad_rows), "failed_rows": len(bad_rows), "status": "invalid" if errors else "ready"}


def error_workbook(errors: list[dict[str, Any]]) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "错误明细"
    sheet.append(["行号", "字段", "错误原因", "原值", "修正建议"])
    for item in errors:
        sheet.append([item["row"], item["column"], item["message"], _excel_value(item.get("value")), "修正后重新上传并从头校验"])
    return _save(book)


def export_workbook_stream(
    rows: Iterable[Mapping[str, Any]], metadata: Mapping[str, Any]
) -> ExportArtifact:
    book = Workbook(write_only=True)
    sheet = book.create_sheet("导出数据")
    iterator = iter(rows)
    first = next(iterator, None)
    keys = list(first) if first is not None else ["result"]
    sheet.append(keys)
    row_count = 0
    for row in chain([first], iterator) if first is not None else ():
        sheet.append([_excel_value(row.get(key)) for key in keys])
        row_count += 1
    guide = book.create_sheet("导出说明")
    for key, value in {**metadata, "row_count": row_count}.items():
        guide.append([key, str(value)])
    return ExportArtifact(_save(book), row_count)


def export_workbook(rows: list[dict[str, Any]], metadata: dict[str, Any]) -> bytes:
    return export_workbook_stream(rows, metadata).content


def export_pdf_stream(
    rows: Iterable[Mapping[str, Any]], metadata: Mapping[str, Any]
) -> ExportArtifact:
    output = BytesIO()
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    document = canvas.Canvas(output)
    document.setTitle(f"ADP data export - {metadata['resource']}")
    document.setAuthor(str(metadata.get("actor") or "ADP"))
    document.setFont("STSong-Light", 12)
    y = 800
    document.drawString(40, y, f"ADP 企业数据导出 · {metadata['resource']}")
    for key, value in metadata.items():
        y -= 20
        document.drawString(40, y, f"{key}: {value}")
    row_count = 0
    for row_count, row in enumerate(rows, start=1):
        y -= 20
        if y < 50:
            document.showPage()
            document.setFont("STSong-Light", 10)
            y = 800
        text = "; ".join(f"{key}={value}" for key, value in row.items())
        lines = wrap(f"{row_count}. {text}", width=110, break_long_words=False, break_on_hyphens=False) or [""]
        for line_index, line in enumerate(lines):
            if line_index:
                y -= 14
                if y < 50:
                    document.showPage()
                    document.setFont("STSong-Light", 10)
                    y = 800
            document.drawString(40, y, line)
    y -= 20
    if y < 50:
        document.showPage()
        document.setFont("STSong-Light", 10)
        y = 800
    document.drawString(40, y, f"row_count: {row_count}")
    document.setSubject(f"row_count={row_count}; request_id={metadata.get('request_id', '')}")
    document.setKeywords(f"resource={metadata['resource']}")
    document.save()
    return ExportArtifact(output.getvalue(), row_count)


def export_pdf(rows: list[dict[str, Any]], metadata: dict[str, Any]) -> bytes:
    return export_pdf_stream(rows, metadata).content


def _save(book: Workbook) -> bytes:
    output = BytesIO()
    book.save(output)
    return output.getvalue()


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _excel_value(value: Any) -> Any:
    value = _json_value(value)
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _validate(required: bool, kind: str, allowed: tuple[str, ...], value: Any) -> str | None:
    if value in (None, ""):
        return "必填项不能为空" if required else None
    if allowed and str(value) not in allowed:
        return f"仅允许：{'/'.join(allowed)}"
    try:
        if kind in {"positive", "number"} and (number := float(value)) < (0 if kind == "number" else 0.0000001):
            return "必须是正数" if kind == "positive" else "不能为负数"
        if kind == "integer" and (int(value) != float(value) or int(value) < 1):
            return "必须是正整数"
        if kind == "date" and isinstance(value, str):
            date.fromisoformat(value)
        if kind == "datetime":
            if isinstance(value, datetime):
                return None
            text = str(value).strip()
            if "T" not in text and " " not in text:
                raise ValueError
            datetime.fromisoformat(text)
    except (TypeError, ValueError):
        return {"positive": "必须是正数", "number": "必须是数字", "integer": "必须是正整数", "date": "日期格式必须为 YYYY-MM-DD", "datetime": "日期时间格式必须为 YYYY-MM-DD HH:MM[:SS]"}.get(kind, "格式不正确")
    return None
